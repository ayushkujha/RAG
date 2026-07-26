import os
import re
import fitz  # PyMuPDF
import chromadb
from chromadb.utils import embedding_functions
from dotenv import load_dotenv
from pypdf import PdfReader
from google import genai
from google.genai import types

# Load environment variables automatically
load_dotenv()

class RAGEngine:
    def __init__(self, api_key=None, db_path="./chroma_db"):
        self.api_key = api_key or os.environ.get("GEMINI_API_KEY")
        self.client = None
        if self.api_key:
            try:
                self.client = genai.Client(api_key=self.api_key)
            except Exception:
                self.client = None
            
        # Initialize persistent ChromaDB Vector Store
        self.db_path = db_path
        self.chroma_client = chromadb.PersistentClient(path=self.db_path)
        self.default_ef = embedding_functions.DefaultEmbeddingFunction()
        
        self.collection = self.chroma_client.get_or_create_collection(
            name="pdf_documents",
            metadata={"hnsw:space": "cosine"}
        )

    def set_api_key(self, api_key):
        """Dynamically update API key and GenAI client."""
        self.api_key = api_key
        if api_key:
            os.environ["GEMINI_API_KEY"] = api_key
            try:
                self.client = genai.Client(api_key=api_key)
            except Exception:
                self.client = None

    def extract_text_from_pdf(self, file_path):
        """Extract text from PDF page by page using PyMuPDF (fitz), fallback to pypdf."""
        pages_content = []
        try:
            doc = fitz.open(file_path)
            for idx, page in enumerate(doc):
                text = page.get_text("text")
                if text and text.strip():
                    pages_content.append({"text": text, "page": idx + 1})
            doc.close()
        except Exception:
            try:
                reader = PdfReader(file_path)
                for idx, page in enumerate(reader.pages):
                    text = page.extract_text()
                    if text and text.strip():
                        pages_content.append({"text": text, "page": idx + 1})
            except Exception:
                pass
                    
        return pages_content

    def extract_text_from_docx(self, file_path):
        """Extract text from Word document (.docx)."""
        import docx
        doc = docx.Document(file_path)
        paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
        return [{"text": "\n".join(paragraphs), "page": 1}]

    def extract_text_from_csv(self, file_path):
        """Extract rows from CSV file."""
        import csv
        rows = []
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            reader = csv.reader(f)
            for row in reader:
                if any(row):
                    rows.append(" | ".join([cell.strip() for cell in row]))
        return [{"text": "\n".join(rows), "page": 1}]

    def extract_text_from_xlsx(self, file_path):
        """Extract rows from Excel workbook (.xlsx)."""
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheets_content = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_rows = []
            for row in sheet.iter_rows(values_only=True):
                if any(row):
                    sheet_rows.append(" | ".join([str(val).strip() if val is not None else "" for val in row]))
            if sheet_rows:
                sheets_content.append(f"### Sheet: {sheet_name}\n" + "\n".join(sheet_rows))
        return [{"text": "\n\n".join(sheets_content), "page": 1}]

    def parse_document(self, file_path):
        """Universal document extractor for PDF, DOCX, CSV, XLSX, TXT, and MD files."""
        ext = os.path.splitext(file_path)[1].lower()
        if ext == ".pdf":
            return self.extract_text_from_pdf(file_path)
        elif ext == ".docx":
            return self.extract_text_from_docx(file_path)
        elif ext == ".csv":
            return self.extract_text_from_csv(file_path)
        elif ext in [".xlsx", ".xls"]:
            return self.extract_text_from_xlsx(file_path)
        else:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                content = f.read()
            return [{"text": content, "page": 1}]

    def chunk_text(self, text_blocks, chunk_size=800, chunk_overlap=150):
        """Split text blocks into overlapping chunks with page metadata."""
        chunks = []
        for block in text_blocks:
            text = block["text"]
            page = block["page"]
            text = re.sub(r'\s+', ' ', text).strip()
            
            start = 0
            while start < len(text):
                end = start + chunk_size
                chunk_text = text[start:end]
                
                chunks.append({
                    "text": chunk_text,
                    "metadata": {
                        "page": page,
                        "char_start": start,
                        "char_end": min(end, len(text))
                    }
                })
                start += (chunk_size - chunk_overlap)
                
        return chunks

    def _local_embedding(self, text):
        """Instant zero-dependency 384-dimensional term vectorizer for keyless execution."""
        if isinstance(text, list):
            return [self._local_embedding(t) for t in text]
            
        import numpy as np
        vec = np.zeros(384, dtype=np.float32)
        words = re.findall(r'\w+', str(text).lower())
        if words:
            for word in words:
                idx = abs(hash(word)) % 384
                vec[idx] += 1.0
            norm = np.linalg.norm(vec)
            if norm > 0:
                vec = vec / norm
        return vec.tolist()

    def get_embedding(self, text):
        """Generate vector embeddings via Gemini API if available, otherwise use instant local vectorizer."""
        key = self.api_key or os.environ.get("GEMINI_API_KEY")
        if key:
            if not self.client:
                try:
                    self.client = genai.Client(api_key=key)
                except Exception:
                    self.client = None

            if self.client:
                embedding_models = ["gemini-embedding-001", "gemini-embedding-2", "gemini-embedding-2-preview"]
                contents = text if isinstance(text, list) else [text]
                for model_name in embedding_models:
                    try:
                        response = self.client.models.embed_content(
                            model=model_name,
                            contents=contents
                        )
                        if response.embeddings:
                            return [emb.values for emb in response.embeddings] if isinstance(text, list) else response.embeddings[0].values
                        elif hasattr(response, 'embedding') and response.embedding:
                            return [emb.values for emb in response.embedding] if isinstance(text, list) else response.embedding.values
                    except Exception:
                        continue

        return self._local_embedding(text)

    def add_document(self, file_name, file_path, chunk_size=800, chunk_overlap=150):
        """Extract text, chunk, embed, and store document in ChromaDB."""
        text_blocks = self.parse_document(file_path)
        chunks = self.chunk_text(text_blocks, chunk_size, chunk_overlap)
        if not chunks:
            return 0

        batch_size = 50
        total_chunks = len(chunks)
        
        for idx in range(0, total_chunks, batch_size):
            batch = chunks[idx:idx + batch_size]
            batch_texts = [c["text"] for c in batch]
            
            embeddings = self.get_embedding(batch_texts)
            
            ids = []
            metadatas = []
            documents = []
            
            for i, chunk in enumerate(batch):
                global_idx = idx + i
                chunk_id = f"{file_name}_chunk_{global_idx}"
                ids.append(chunk_id)
                documents.append(chunk["text"])
                metadatas.append({
                    "source": file_name,
                    "page": int(chunk["metadata"]["page"]),
                    "chunk_index": int(global_idx)
                })
                
            self.collection.add(
                ids=ids,
                embeddings=embeddings,
                metadatas=metadatas,
                documents=documents
            )
            
        return total_chunks

    def get_indexed_documents(self):
        """Retrieve list of indexed unique document filenames and total chunk count from ChromaDB."""
        data = self.collection.get(include=["metadatas"])
        metadatas = data.get("metadatas", [])
        
        doc_counts = {}
        for meta in metadatas:
            src = meta.get("source", "Unknown")
            doc_counts[src] = doc_counts.get(src, 0) + 1
            
        return doc_counts

    def count_chunks(self):
        """Return total number of chunks stored in ChromaDB."""
        return self.collection.count()

    def delete_document(self, file_name):
        """Delete all chunks belonging to a specific document from ChromaDB."""
        self.collection.delete(where={"source": file_name})
        return True

    def clear_index(self):
        """Reset the ChromaDB collection completely."""
        self.chroma_client.delete_collection("pdf_documents")
        self.collection = self.chroma_client.get_or_create_collection(
            name="pdf_documents",
            metadata={"hnsw:space": "cosine"}
        )
        return True

    def retrieve(self, query, top_k=4):
        """Query ChromaDB using semantic embedding search."""
        if self.collection.count() == 0:
            return []

        query_embedding = self.get_embedding(query)
        results = self.collection.query(
            query_embeddings=[query_embedding],
            n_results=min(top_k, self.collection.count()),
            include=["documents", "metadatas", "distances"]
        )

        chunks = []
        if results and results["documents"] and len(results["documents"]) > 0:
            docs = results["documents"][0]
            metas = results["metadatas"][0]
            distances = results["distances"][0] if "distances" in results and results["distances"] else [0]*len(docs)
            
            for doc_text, meta, dist in zip(docs, metas, distances):
                similarity = 1.0 - dist if dist <= 1.0 else (1.0 / (1.0 + dist))
                chunks.append({
                    "text": doc_text,
                    "metadata": meta,
                    "similarity": similarity
                })

        return chunks

    def query_with_context(self, query, chat_history=None, top_k=4):
        """Retrieve context from ChromaDB and generate an answer using Gemini or local context synthesis."""
        retrieved_chunks = self.retrieve(query, top_k=top_k)
        
        if retrieved_chunks:
            context_blocks = []
            for chunk in retrieved_chunks:
                source = chunk["metadata"].get("source", "Document")
                page = chunk["metadata"].get("page", 1)
                context_blocks.append(f"[Source: {source}, Page: {page}]\n{chunk['text']}")
            context_str = "\n\n".join(context_blocks)
        else:
            context_str = "No relevant document context found in uploaded files."

        key = self.api_key or os.environ.get("GEMINI_API_KEY")
        if key:
            if not self.client:
                try:
                    self.client = genai.Client(api_key=key)
                except Exception:
                    self.client = None

            if self.client:
                system_instruction = (
                    "You are an AI Document Assistant. You answer user questions strictly based on the provided context snippets.\n"
                    "If the context does not contain enough information to answer the question, state politely that the uploaded documents do not contain that information.\n"
                    "Always cite the relevant document name and page number when answering."
                )

                current_prompt = (
                    f"Context from uploaded documents:\n{context_str}\n\n"
                    f"User Question: {query}"
                )

                messages = []
                if chat_history:
                    for msg in chat_history:
                        role = "user" if msg["role"] == "user" else "model"
                        messages.append(types.Content(
                            role=role,
                            parts=[types.Part.from_text(text=msg["content"])]
                        ))
                        
                messages.append(types.Content(
                    role="user",
                    parts=[types.Part.from_text(text=current_prompt)]
                ))

                try:
                    response = self.client.models.generate_content(
                        model="gemini-2.5-flash",
                        contents=messages,
                        config=types.GenerateContentConfig(
                            system_instruction=system_instruction,
                            temperature=0.2,
                        )
                    )
                    return response.text, retrieved_chunks
                except Exception as e:
                    print(f"GenAI call error: {e}")

        # Natural response synthesis from retrieved document context
        if retrieved_chunks:
            top_contexts = []
            for c in retrieved_chunks[:2]:
                src = c["metadata"].get("source", "Document")
                pg = c["metadata"].get("page", 1)
                top_contexts.append(f"**From {src} (Page {pg}):**\n\n{c['text']}")
            response_text = "\n\n---\n\n".join(top_contexts)
        else:
            response_text = "I couldn't find any relevant information in your uploaded documents regarding that question."

        return response_text, retrieved_chunks
